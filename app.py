import os
import json
import numpy as np
import pandas as pd
import streamlit as st
import torch
import yaml
from openpyxl import workbook  # pip install openpyxl
from openpyxl import load_workbook
from PIL import Image
from sklearn.neighbors import KNeighborsRegressor
import time
from commons.json_tools import get_data_from_json
from constants import CLASS_FOLDER
from segmentation.predict_onnx import predict_onnx, predict
from segmentation.segment_sam import segment
from yolo.constants import CONFIG_FOLDER, RUNS_FOLDER
from hubconf import custom  

# Obtenir les noms des classes
def get_class_names(yaml_file):
    with open(yaml_file, errors="ignore") as f:
        data = yaml.safe_load(f)  # dictionary
    return data["names"]

# Récuperer tous les fichiers jpg dans un dossier
def get_jpg_files(files):
    jpg_files = [file for file in files if file.lower().endswith(".jpg") or file.lower().endswith(".png")]
    return jpg_files

# Charger les model de segmentation unet
@st.cache_resource(show_spinner=False)
def get_seg_model(device):
    model = torch.load(f="segmentation/unet16_model.pt", map_location=device )
    return model

# Calculer le nombre et le poids des objets détectés
def detect_count_objects(_results, images, model, model_type, conf_thres=[], yaml_file="", data_path="", detection_type="weight", k_neighbors=10, field=1, device="cpu"):
    class_names = get_class_names(yaml_file)
    dict_detect = {class_names[class_detect]: {"quantity": 0, "area": 0} for class_detect in range(len(class_names))}
    img_crv = []
    areas_crv = []
    hgts = []
    wdts = []
    crv_knn_weight = 0
    areas_i = np.array([])
    heights_i = np.array([])
    widths_i = np.array([])
    # Parcourir les images et les objets détectés
    for result, img in zip(_results.pandas().xyxy, images):
        boxs = []
        # yolo format - (class_id, x_center, y_center, width, height)
        # coco format - (annotation_id, x_upper_left, y_upper_left, width, height)
        img_w, img_h = img.size
        for i in range(len(result)):
            class_detect = result.loc[i, "class"]
            confidence = result.loc[i, "confidence"]
            xmin = result.loc[i, "xmin"]
            ymin = result.loc[i, "ymin"]
            xmax = result.loc[i, "xmax"]
            ymax = result.loc[i, "ymax"]
            width = (xmax - xmin) / img_w
            height = (ymax - ymin) / img_h
            area = float(width) * float(height)
            # Vérifier que la confiance est supérieur au seuil
            if confidence >= conf_thres[class_detect]:
                dict_detect[class_names[class_detect]]["quantity"] += 1
                # Utilisation du KNN pour la classe conserve ronde
                if class_names[class_detect] == "conserve_ronde":
                    if detection_type == "area" or detection_type == "knn":
                        # Crop des objets détectés
                        boxs.append([xmin, ymin, xmax, ymax])
                        cropped_img = img.crop((xmin, ymin, xmax, ymax))
                        cropped_img = cropped_img.resize((256, 256))
                        # cropped_img.save("img_"+str(i)+".jpg")
                        # i+=1
                        img_crv.append(cropped_img)
                        areas_crv.append(area)
                        if detection_type == "knn":
                            hgts.append(max(height, width))
                            wdts.append(min(width, height))
                    else:
                        dict_detect[class_names[class_detect]]["area"] += area

                else:
                    dict_detect[class_names[class_detect]]["area"] += area
        a_, h_, w_ = segment(img,boxs,device)
        areas_i = np.concatenate((areas_i,a_))
        heights_i = np.concatenate((heights_i,h_))
        widths_i = np.concatenate((widths_i,w_))
    if detection_type == "area" or detection_type == "knn":
        # Utilisation du modèle ONNX pour la segmentation des objets détectés
        if model_type == 'onnx':
            areas, heights, widths = predict_onnx(model_path="segmentation/unet16_model.onnx", images=img_crv, areas=areas_crv, hgts=hgts, wdts=wdts, device=device)
        # Utilisation du modèle Pytorch pour la segmentation des objets détectés
        else :
            #areas, heights, widths = predict(model=model, images=img_crv, areas=areas_crv, hgts=hgts, wdts=wdts, device=device)
            areas_sam, heights_sam, widths_sam = areas_i, heights_i, widths_i
            # area_comp = dict()
            # area_comp["sam"] = areas_sam
            # area_comp["old"] = areas
            # area_data = pd.DataFrame.from_dict(area_comp)
            # area_data.to_csv("./area_data.csv")
        dict_detect["conserve_ronde"]["area"] = np.sum(areas_sam)
        # Application du knn
        if detection_type == "knn":
            #crv_knn_weight = detect_weight_knn(data_path, areas, heights, widths, k_neighbors, field)
            crv_knn_weight = detect_weight_knn(data_path, areas_sam, heights_sam, widths_sam, k_neighbors, field)
    return dict_detect, crv_knn_weight

# Fonction de détection des poids avec knn
def detect_weight_knn(data_path, areas, heights, widths, k_neighbors=10, field=1):
    # Récupérer les données des conserves rondes de réference (poids, longueur, largeur, surface)
    excel_data = pd.read_csv(data_path, delimiter=";")
    data = pd.DataFrame(excel_data, columns=["poids", "h", "w", "S_seg_autre"])
    # Calculer le rapport h/w
    data["rapp_h_w"] = data["h"] / data["w"]
    #data["rapp_h_w"] = (data["rapp_h_w"] - data["rapp_h_w"].mean()) / data["rapp_h_w"].std()
    # Adaptation de l'unité de surface
    data["S_seg_autre"] = data["S_seg_autre"] / 1000000
    #data["S_seg_autre"] = (data["S_seg_autre"] - data["S_seg_autre"].mean()) / data["S_seg_autre"].std()
    weights = data.poids.to_numpy()
    #X = np.array([data.rapp_h_w.to_numpy(), data.S_seg_autre.to_numpy()]).reshape(-1, 2)
    X = np.array([ data.S_seg_autre.to_numpy()]).reshape(-1,1)
    neigh = KNeighborsRegressor(n_neighbors=int(k_neighbors))
    neigh.fit(X, weights)
    # Normaliser la surface
    #areas_n = (field * areas - data["S_seg_autre"].mean()) / data["S_seg_autre"].std()
    areas_n = field*areas
    rapp_h_l = heights / widths
    # Normaliser le rapport h/w
    #rapp_h_l_n = (rapp_h_l - data["rapp_h_w"].mean()) / data["rapp_h_w"].std()
    rapp_h_l_n = rapp_h_l
    data_last_format = [[axe1, axe2] for axe1, axe2 in zip(rapp_h_l_n, areas_n)]
    # Pédire les poids
    weights_hat = neigh.predict(areas_n.reshape(-1,1))
    #area_comp = dict()
    #area_comp["areas"] = areas
    #area_comp["knn_poids"] = weights_hat.tolist()
    #area_data = pd.DataFrame.from_dict(area_comp)
    #area_data.to_csv("./area_weight_"+seg_type+".csv")
    total_weight = np.sum(weights_hat)
    return total_weight

# Calculer les poids des objets détectés en utilisant le poids moyen statistique
def detection_to_weight(count_detect, flow):
    avg_weight_data = get_data_from_json(os.path.join(CLASS_FOLDER, flow, "average_weight.json"))

    weight_detect = {
        avg_weight_data[_class]["name"]: {
            "weight(g)": int(detection["quantity"] * avg_weight_data[_class]["average weight(g)"]),
            "quantity": detection["quantity"],
        }
        for _class, detection in count_detect.items()
    }
    return weight_detect

# Calculer les poids des objets détectés en utilisant la surface moyenne statistique
def detection_to_area(count_detect, flow):
    avg_area_data = get_data_from_json(os.path.join(CLASS_FOLDER, flow, "average_weight.json"))
    area_detect = {
        avg_area_data[_class]["name"]: {
            "weight(norm)": detection["area"] * avg_area_data[_class]["average weight per area(g/cm2)"],
            "area": detection["area"],
        }
        for _class, detection in count_detect.items()
    }
    return area_detect

# Rassembler le nombre de detections et leurs poids dans une structure de données 
def compare_detect_to_manual(classes, weight_detect, area_detect, image_size_cm=100, crv_knn_weight=0):
    coeff_norm_to_m = image_size_cm * 10000
    compare_list = [
        {
            "name": name,
            "detect_qty": weight_detect.get(name)["quantity"] if weight_detect.get(name) else None,
            "detect_weight(g)": weight_detect.get(name)["weight(g)"] if weight_detect.get(name) else None,
            "detect_weight(g)_area": int(area_detect.get(name)["weight(norm)"] * coeff_norm_to_m)
            if area_detect.get(name)
            else None,
            "detect_weight(g)_knn": crv_knn_weight if name == "conserve_ronde" else 0,
        }
        for name in classes
    ]
    return compare_list

# Charger le modèle de détection yolo
@st.cache_resource(show_spinner=False)
def get_yolo_model(flow, exp,model_type, device):
    weight_pathfile = os.path.join(RUNS_FOLDER, flow, "train", exp, "weights", "best." + model_type)
    model = custom(path=weight_pathfile, autoshape=True, _verbose=True, device=device)
    #model = torch.hub.load('ultralytics/yolov5', 'custom', weight_pathfile, device=device)
    return model

# Détecter les objets sur les images avec yolo
@st.cache_data
def get_detection(flow, images, _model_yolo, _model_seg, model_type, detection_type, img_size, device, conf_thres, config_file, k_neighbors, field):
    yaml_file = os.path.join(CONFIG_FOLDER, flow, config_file)
    data_path = os.path.join(CLASS_FOLDER, flow, "db.csv")
    classes = get_class_names(yaml_file)
    read_images = [Image.open(img) for img in images]
    # Inference
    results = _model_yolo(read_images, size=img_size)
    # Calcul de nombre et poids des objets 
    count_detect = detect_count_objects(results, read_images, _model_seg, model_type, conf_thres, yaml_file, data_path, detection_type, k_neighbors, field, device)
    area_detect = detection_to_area(count_detect[0], flow)
    weight_detect = detection_to_weight(count_detect[0], flow)
    # Rassembler les résultats dans une liste
    compare_list = compare_detect_to_manual(
        classes, weight_detect, area_detect, image_size_cm=field, crv_knn_weight=count_detect[1]
    )
    return compare_list

# Générer le rapport de la caractérisation en se basant sur un template
def generate_report(carac_results, save_dir, flow, detection_type):
    # Choix du template qui dépend du flux
    flow_template = 5 if flow=="acier" else 6
    emr_classes_file = os.path.join(CLASS_FOLDER, flow, "classes_emr.json")
    # Charger les classes emr
    with open(emr_classes_file) as f:
            emr_classes = json.load(f)
    classes_file = os.path.join(CLASS_FOLDER, flow, "classes.txt")
    template_file = os.path.join(CLASS_FOLDER, flow, "template.xlsx")
    class2idx = dict()
    with open(classes_file, "r") as f:
        lines = f.read().splitlines()
        for i, line in enumerate(lines):
            class2idx[line] = i
    # Charger le template
    wb = load_workbook(template_file)
    sheets = wb.sheetnames
    worksheet = wb[sheets[0]]
    #poids_tot = st.number_input("Insert total weight (g)")
    #worksheet.cell(row=1, column=2).value = poids_tot
    # Obtenir les poids de la carac manuelle
    carac_weights = dict()
    for k, v in emr_classes.items():
        carac_weights[v] = st.number_input("Insert " + k + " weight (g)")
    # Saisir les poids de la détection dans le template
    for i in range(1, len(carac_results)+1):
        cls = str(worksheet.cell(row=i + 15, column=1).value)
        poids = carac_results[i-1]["detect_weight(g)"]
        if detection_type == "knn" and cls == "conserve_ronde":
            poids = carac_results[class2idx[cls]]["detect_weight(g)_knn"]
        worksheet.cell(row=i + 15, column=2).value = poids
    # Saisir les poids de la carac manuelle dans le template
    for i in range(1, len(emr_classes.keys())+1):
        worksheet.cell(row=i + 15, column=flow_template).value = carac_weights[i - 1]
    # Sauvegarder le rapport
    wb.save(os.path.join(save_dir, "report.xlsx"))
    # Button pour télécharger le rapport
    with open(os.path.join(save_dir, "report.xlsx"), "rb") as f:
        st.download_button("Download report", f, file_name="report.xlsx")

def main():
    # Titre
    st.title("Caracterization using YOLO")
    # Charger le fichier config
    yaml_file = "streamlit_conf.yaml"
    with open(yaml_file, errors="ignore") as f:
        data = yaml.safe_load(f)  # dictionary
    config_file = data["config_file"]
    img_size = data["image_size"]
    save_dir = data["save_dir"]
    # Case pour selectionner le flux
    flow = st.selectbox(
    'Flow ?',
    ('acier', 'alu'))
    # Case pour selectionner le device sur lequel la détection tourne
    select_device = st.selectbox(
    'Device ?',
    ('cuda:0', 'cpu'))
    # Case pour selectionner le modèle yolo et le paramètre k du knn
    if flow=="acier":
        exp = st.selectbox(
        'experience ?',
        ('2022_11_02_exp', 'test_evolve_exp'))
        k_neighbors = st.number_input(
        'Number of neighbors for knn ?',min_value = 1, max_value = 100)
    else :
        exp = 'test_evolve_exp'
        k_neighbors=None
    # Charger les seuils de confiance pour chaque classes
    f1_tresh_file = os.path.join(RUNS_FOLDER, flow, "train", exp, "f1_tresh.yaml")
    with open(f1_tresh_file, errors="ignore") as f:
        data = yaml.safe_load(f)
    conf_thres = data["seuils"]
    device = torch.device(select_device if torch.cuda.is_available() else "cpu")
    # Choix du type de modèle en fonction de du device
    model_type = 'pt' if select_device == "cuda:0" else 'onnx'
    # Choix du champs de vision (field) en fonction du flux (à modifier pendant le calibrage)
    field = 2.614 if flow=="acier" else 1
    # Choix de type de détection en fonction du flux
    detection_type = "knn" if flow=="acier" else "weight"
    # Choix de source de données (local ou upload)
    data_source = st.radio(
    "Data source ? ",
    ('Upload', 'Local'))
    # Case pour charger les images
    if data_source == 'Upload':
        images = st.file_uploader("Import images", type=["png", "jpg", "JPG"], accept_multiple_files=True)
    else :
        images_path = os.path.join(CLASS_FOLDER, flow, "streamlit_input")
        images = [os.path.join(images_path, img) for img in get_jpg_files(os.listdir(images_path))]
    # Lancer la détection quand les images sont chargées
    if len(images):
        data_load_state = st.text("Detecting...")
        model_yolo=get_yolo_model(flow, exp, model_type, device)
        if model_type=='onnx':
            model_seg=None
        else :
            model_seg=get_seg_model(device)
        starttime = time.time()
        compare_list = get_detection(
            flow, images, model_yolo, model_seg, model_type, detection_type, img_size, device, conf_thres, config_file, k_neighbors, field
        )
        # Calcul du temps de détection en seconde
        curr_time = time.time() - starttime
        st.text('Detection time : '+ str(curr_time) + ' s' )
        data_load_state.text("Detecting...done!")
        st.header("Caracterization results : ")
        # Afficher le tableau des résultats de la détection
        st.table(compare_list)
        # Générer le rapport
        generate_report(compare_list, save_dir, flow, detection_type)

# Fonction d'authentification
def check_password():
    """Returns `True` if the user had a correct password."""

    def password_entered():
        """Checks whether a password entered by the user is correct."""
        if (
            st.session_state["username"] in st.secrets["passwords"]
            and st.session_state["password"]
            == st.secrets["passwords"][st.session_state["username"]]
        ):
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # don't store username + password
            del st.session_state["username"]
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        # First run, show inputs for username + password.
        st.text_input("Username", on_change=password_entered, key="username")
        st.text_input(
            "Password", type="password", on_change=password_entered, key="password"
        )
        return False
    elif not st.session_state["password_correct"]:
        # Password not correct, show input + error.
        st.text_input("Username", on_change=password_entered, key="username")
        st.text_input(
            "Password", type="password", on_change=password_entered, key="password"
        )
        st.error("😕 User not known or password incorrect")
        return False
    else:
        # Password correct.
        return True

if __name__ == "__main__":
    if check_password():
        main()
